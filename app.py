"""
Torre de Controle de Estoque, Quarentena e Cobertura
Filial CDSC | Deploy: Streamlit Cloud
"""

import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURAÇÕES GLOBAIS
# Ajuste aqui se os nomes de abas ou colunas mudarem
# ═══════════════════════════════════════════════════════════════════════════════

NOME_ABA_DEMANDA     = "CDSC - BLOQ ITEM"
NOME_ABA_QUARENTENA  = "DW_ESTOQUE_QUARENTENA (2)"
NOME_ABA_PLANEJADO   = "PLAN VENDA VS VENDIDO - CDSC"
NOME_ABA_ATENDIMENTO = "CDSC"

LINHA_CABECALHO_DEMANDA = 22  # cabeçalho na linha 22, dados a partir da 23

COLUNAS_DEMANDA = {
    "codigo":             "Produto Cod",
    "descricao":          "Produto Desc",
    "qtd_programada":     "Qtd Programada",
    "estoque_disponivel": "Estoque Disponivel",
    "saldo":              "Saldo",
    "status":             "Status",
}

COLUNAS_IGNORAR_DEMANDA = ["Quarentena + Saldo", "Data Chegada"]

COLUNAS_QUARENTENA = {
    "codigo":           "SKU",            # Troque por "Produto Cod" se necessário
    "lote":             "Lote",
    "descricao":        "Descrição",
    "quantidade":       "Qtde",           # Troque por "Total" se necessário
    "data_fabricacao":  "Data Fabricação",
    "dias_antecipacao": "Dias Antecipação",
    "liberacao":        "Liberação",
}

COLUNAS_PLANEJADO = {
    "codigo":      "Produto Cod",
    "cota":        "Cota",
    "qtd_liq_lib": "Qtde Líq + Lib",
    "diferenca":   "Diferença Qtde Líq+Lib",
    "percentual":  "Percentual Qtde Líq+Lib",
}

COLUNAS_ATENDIMENTO = {
    "codigo":             "Produto Cod",
    "descricao":          "Produto Desc",
    "producao_prevista":  "Produção Prevista",
    "data_producao":      "Data Produção",
    "cobertura_futura":   "Cobertura Futura",
    "status_cobertura":   "Status Cobertura",
    "observacao":         "Observação Planejamento",
}

CORES_STATUS_EXCEL = {
    "COBERTURA VIA QUARENTENA": "00B050",
    "COBERTURA PARCIAL":        "FFFF00",
    "COBERTURA FUTURA":         "00B0F0",
    "AGUARDAR PRODUÇÃO":        "FF6600",
    "SEM PRODUÇÃO NO MÊS":      "BFBFBF",
    "CRÍTICO":                  "FF0000",
}

CORES_STATUS_CSS = {
    "COBERTURA VIA QUARENTENA": "background-color:#c6efce;color:#276221;font-weight:bold",
    "COBERTURA PARCIAL":        "background-color:#ffeb9c;color:#9c6500;font-weight:bold",
    "COBERTURA FUTURA":         "background-color:#bdd7ee;color:#1f497d;font-weight:bold",
    "AGUARDAR PRODUÇÃO":        "background-color:#fcd5b4;color:#974706;font-weight:bold",
    "SEM PRODUÇÃO NO MÊS":      "background-color:#e0e0e0;color:#444444;font-weight:bold",
    "CRÍTICO":                  "background-color:#ffc7ce;color:#9c0006;font-weight:bold",
}

# ═══════════════════════════════════════════════════════════════════════════════
# UTILITÁRIOS
# ═══════════════════════════════════════════════════════════════════════════════

def padronizar_codigo(valor) -> str:
    """Converte código para string limpa sem decimais ou espaços."""
    if valor is None:
        return ""
    if isinstance(valor, float) and np.isnan(valor):
        return ""
    s = str(valor).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.strip()


def padronizar_data(valor) -> str:
    """Converte data para dd/mm/yyyy."""
    try:
        if pd.isnull(valor) or valor == "":
            return ""
    except Exception:
        pass
    try:
        return pd.to_datetime(valor, dayfirst=True).strftime("%d/%m/%Y")
    except Exception:
        return str(valor) if valor else ""


def padronizar_numero(valor) -> float:
    """Converte para float, retorna 0.0 se vazio ou inválido."""
    if valor is None:
        return 0.0
    if isinstance(valor, float) and np.isnan(valor):
        return 0.0
    try:
        if isinstance(valor, str):
            valor = valor.strip().replace(".", "").replace(",", ".")
        return float(valor)
    except Exception:
        return 0.0


def _renomear_colunas(df: pd.DataFrame, mapa: dict) -> pd.DataFrame:
    """Renomeia colunas do df usando mapeamento {chave_interna: nome_original}."""
    inverso = {v: k for k, v in mapa.items()}
    df = df.rename(columns=inverso)
    for chave in mapa:
        if chave not in df.columns:
            df[chave] = None
    return df


def _limpar_sem_codigo(df: pd.DataFrame) -> pd.DataFrame:
    """Remove linhas sem código de produto."""
    df = df[df["codigo"].notna()].copy()
    df["codigo"] = df["codigo"].apply(padronizar_codigo)
    df = df[df["codigo"] != ""]
    return df.reset_index(drop=True)

# ═══════════════════════════════════════════════════════════════════════════════
# LEITURA DAS BASES
# ═══════════════════════════════════════════════════════════════════════════════

def read_demanda(file) -> pd.DataFrame:
    """Lê aba CDSC - BLOQ ITEM a partir da linha 22 como cabeçalho."""
    try:
        df = pd.read_excel(
            file,
            sheet_name=NOME_ABA_DEMANDA,
            header=LINHA_CABECALHO_DEMANDA - 1,
            dtype=str,
        )
        for col in COLUNAS_IGNORAR_DEMANDA:
            if col in df.columns:
                df.drop(columns=[col], inplace=True)
        df = _renomear_colunas(df, COLUNAS_DEMANDA)
        df = _limpar_sem_codigo(df)
        df["qtd_programada"]     = df["qtd_programada"].apply(padronizar_numero)
        df["estoque_disponivel"] = df["estoque_disponivel"].apply(padronizar_numero)
        df["saldo"]              = df["saldo"].apply(padronizar_numero)
        return df
    except Exception as e:
        st.error(f"❌ Erro ao ler Análise de Demanda: {e}")
        return pd.DataFrame()


def read_quarentena(file) -> pd.DataFrame:
    """Lê aba DW_ESTOQUE_QUARENTENA (2)."""
    try:
        df = pd.read_excel(file, sheet_name=NOME_ABA_QUARENTENA, dtype=str)
        df = _renomear_colunas(df, COLUNAS_QUARENTENA)
        df = _limpar_sem_codigo(df)
        df["quantidade"] = df["quantidade"].apply(padronizar_numero)
        df["liberacao"]  = df["liberacao"].apply(padronizar_data)
        return df
    except Exception as e:
        st.error(f"❌ Erro ao ler DW Quarentena: {e}")
        return pd.DataFrame()


def read_planejado_vendido(file) -> pd.DataFrame:
    """Lê aba PLAN VENDA VS VENDIDO - CDSC."""
    try:
        df = pd.read_excel(file, sheet_name=NOME_ABA_PLANEJADO, dtype=str)
        df = _renomear_colunas(df, COLUNAS_PLANEJADO)
        df = _limpar_sem_codigo(df)
        df["cota"]        = df["cota"].apply(padronizar_numero)
        df["qtd_liq_lib"] = df["qtd_liq_lib"].apply(padronizar_numero)
        df["diferenca"]   = df["diferenca"].apply(padronizar_numero)

        def parse_percentual(v):
            if v is None:
                return 0.0
            try:
                return float(str(v).replace("%", "").replace(",", ".").strip())
            except Exception:
                return 0.0

        df["percentual"] = df["percentual"].apply(parse_percentual)
        return df
    except Exception as e:
        st.error(f"❌ Erro ao ler Planejado x Vendido: {e}")
        return pd.DataFrame()


def read_atendimento(file) -> pd.DataFrame:
    """Lê aba CDSC da Planilha de Atendimento."""
    try:
        df = pd.read_excel(file, sheet_name=NOME_ABA_ATENDIMENTO, dtype=str)
        df = _renomear_colunas(df, COLUNAS_ATENDIMENTO)
        df = _limpar_sem_codigo(df)
        df["producao_prevista"] = df["producao_prevista"].apply(padronizar_numero)
        df["data_producao"]     = df["data_producao"].apply(padronizar_data)
        return df
    except Exception as e:
        st.error(f"❌ Erro ao ler Planilha de Atendimento: {e}")
        return pd.DataFrame()

# ═══════════════════════════════════════════════════════════════════════════════
# ANÁLISE
# ═══════════════════════════════════════════════════════════════════════════════

def identificar_itens_faltantes(df_demanda: pd.DataFrame) -> pd.DataFrame:
    """Filtra itens com Saldo < 0 ou Status = Falta de Estoque."""
    mask_saldo  = df_demanda["saldo"] < 0
    mask_status = (
        df_demanda["status"].astype(str).str.strip().str.upper() == "FALTA DE ESTOQUE"
    )
    df = df_demanda[mask_saldo | mask_status].copy()
    df["falta_atual"] = df["saldo"].apply(lambda x: abs(x) if x < 0 else 0)
    return df.reset_index(drop=True)


def calcular_quarentena(df_faltantes: pd.DataFrame, df_quarentena: pd.DataFrame) -> pd.DataFrame:
    """
    Para cada item faltante busca TODOS os lotes no DW_QUARENTENA.
    Soma total e formata próximas liberações agrupadas por data.
    """
    if df_quarentena.empty:
        df_faltantes["quarentena_total"]    = 0.0
        df_faltantes["proximas_liberacoes"] = ""
        return df_faltantes

    def resumo_grupo(grupo):
        grupo = grupo[grupo["quantidade"] > 0]
        if grupo.empty:
            return pd.Series({"quarentena_total": 0.0, "proximas_liberacoes": ""})

        total    = grupo["quantidade"].sum()
        por_data = (
            grupo.groupby("liberacao", dropna=False)["quantidade"]
            .sum()
            .reset_index()
        )
        # Ordena cronologicamente
        try:
            por_data["_dt"] = pd.to_datetime(
                por_data["liberacao"], format="%d/%m/%Y", errors="coerce"
            )
            por_data = por_data.sort_values("_dt")
        except Exception:
            pass

        partes = []
        for _, row in por_data.iterrows():
            data_str = row["liberacao"] if row["liberacao"] not in ("", None, "nan") else "s/data"
            qtd_fmt  = f"{int(row['quantidade']):,}".replace(",", ".")
            partes.append(f"{data_str}: {qtd_fmt} und")

        return pd.Series({
            "quarentena_total":    total,
            "proximas_liberacoes": " | ".join(partes),
        })

    resumo = (
        df_quarentena.groupby("codigo")
        .apply(resumo_grupo)
        .reset_index()
    )

    df_result = df_faltantes.merge(resumo, on="codigo", how="left")
    df_result["quarentena_total"]    = df_result["quarentena_total"].fillna(0.0)
    df_result["proximas_liberacoes"] = df_result["proximas_liberacoes"].fillna("")
    return df_result


def consultar_planejado_vendido(df: pd.DataFrame, df_planejado: pd.DataFrame) -> pd.DataFrame:
    """Adiciona percentual vendido e alerta comercial."""
    if df_planejado.empty:
        df["percentual_vendido"] = 0.0
        df["excesso_vendido"]    = 0.0
        df["alerta_comercial"]   = ""
        return df

    df_p = df_planejado[["codigo", "percentual", "diferenca"]].copy()
    df   = df.merge(df_p, on="codigo", how="left")
    df["percentual"] = df["percentual"].fillna(0.0)
    df["diferenca"]  = df["diferenca"].fillna(0.0)
    df.rename(columns={
        "percentual": "percentual_vendido",
        "diferenca":  "excesso_vendido",
    }, inplace=True)

    df["alerta_comercial"] = df["percentual_vendido"].apply(
        lambda p: f"Venda {p:.0f}% acima do planejado" if (p or 0) > 100 else ""
    )
    return df


def consultar_atendimento(df: pd.DataFrame, df_atendimento: pd.DataFrame) -> pd.DataFrame:
    """Adiciona cobertura futura, produção prevista e observação de planejamento."""
    defaults = {
        "producao_prevista":       0.0,
        "data_producao":           "",
        "cobertura_futura":        "",
        "status_cobertura":        "",
        "observacao_planejamento": "",
    }

    if df_atendimento.empty:
        for col, val in defaults.items():
            df[col] = val
        return df

    cols_usar = [c for c in [
        "codigo", "producao_prevista", "data_producao",
        "cobertura_futura", "status_cobertura", "observacao"
    ] if c in df_atendimento.columns]

    df = df.merge(df_atendimento[cols_usar], on="codigo", how="left")

    for col, default in defaults.items():
        src = "observacao" if col == "observacao_planejamento" else col
        if src in df.columns and src != col:
            df.rename(columns={src: col}, inplace=True)
        if col not in df.columns:
            df[col] = default
        else:
            df[col] = df[col].fillna(default)

    return df


def definir_status_operacional(row) -> str:
    """
    Regras de prioridade:
    1. Tem quarentena suficiente → COBERTURA VIA QUARENTENA
    2. Tem quarentena parcial   → COBERTURA PARCIAL
    3. Tem cobertura futura     → COBERTURA FUTURA
    4. Tem produção prevista    → AGUARDAR PRODUÇÃO
    5. Sem produção no mês      → SEM PRODUÇÃO NO MÊS
    6. Nenhuma das anteriores   → CRÍTICO
    """
    falta   = float(row.get("falta_atual", 0) or 0)
    q       = float(row.get("quarentena_total", 0) or 0)
    prod    = float(row.get("producao_prevista", 0) or 0)
    cob     = str(row.get("cobertura_futura", "") or "").strip()
    st_cob  = str(row.get("status_cobertura", "") or "").strip().upper()
    dt_p    = str(row.get("data_producao", "") or "").strip()

    VAZIOS      = {"", "0", "nan", "0.0", "none", "nat"}
    tem_q       = q > 0
    tem_prod    = prod > 0 or dt_p.lower() not in VAZIOS
    tem_cob_f   = cob.lower() not in VAZIOS
    sem_mes     = "SEM PRODUÇÃO" in st_cob or (not tem_prod and not tem_cob_f)

    if tem_q:
        return "COBERTURA VIA QUARENTENA" if q >= falta else "COBERTURA PARCIAL"
    if tem_cob_f:
        return "COBERTURA FUTURA"
    if tem_prod:
        return "AGUARDAR PRODUÇÃO"
    if sem_mes:
        return "SEM PRODUÇÃO NO MÊS"
    return "CRÍTICO"


def gerar_observacao_sistema(row) -> str:
    """Gera nota automática sobre déficit e pressão comercial."""
    notas = []
    falta = float(row.get("falta_atual", 0) or 0)
    q     = float(row.get("quarentena_total", 0) or 0)

    if 0 < q < falta:
        deficit = int(falta - q)
        notas.append(f"Déficit de {deficit:,} und após quarentena.".replace(",", "."))
    if row.get("alerta_comercial", ""):
        notas.append("Pressão comercial identificada.")
    return " ".join(notas)

# ═══════════════════════════════════════════════════════════════════════════════
# GERAÇÃO DO EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def _formatar_aba(ws, df: pd.DataFrame):
    """Aplica formatação visual na aba."""
    # Cabeçalho
    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill(fill_type="solid", fgColor="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align

    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions

    # Largura das colunas
    for col_idx, col_name in enumerate(df.columns, 1):
        try:
            max_len = df[col_name].astype(str).str.len().max()
            max_len = max(max_len if not np.isnan(max_len) else 0, len(str(col_name)))
        except Exception:
            max_len = len(str(col_name))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 55)

    # Colorir coluna Status Operacional
    if "Status Operacional" in df.columns:
        st_idx = list(df.columns).index("Status Operacional") + 1
        for row_idx in range(2, ws.max_row + 1):
            cell  = ws.cell(row=row_idx, column=st_idx)
            val   = str(cell.value or "")
            cor   = CORES_STATUS_EXCEL.get(val)
            if cor:
                cell.fill = PatternFill(fill_type="solid", fgColor=cor)
                escuro    = val in ("COBERTURA PARCIAL", "COBERTURA VIA QUARENTENA")
                cell.font = Font(bold=True, color="000000" if escuro else "FFFFFF")


def gerar_excel(df_final: pd.DataFrame) -> BytesIO:
    """Gera arquivo Excel com 4 abas e formatação."""
    output = BytesIO()

    df_criticos  = df_final[df_final["Status Operacional"] == "CRÍTICO"].copy()
    df_qcob      = df_final[df_final["Status Operacional"].isin(
                       ["COBERTURA VIA QUARENTENA", "COBERTURA PARCIAL"])].copy()
    df_venda     = df_final[df_final["Alerta Comercial"] != ""].copy()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for df_tab, nome_aba in [
            (df_final,  "Relatório Final"),
            (df_criticos,"Itens Críticos"),
            (df_qcob,   "Cobertura por Quarentena"),
            (df_venda,  "Venda Acima do Planejado"),
        ]:
            df_tab.to_excel(writer, sheet_name=nome_aba, index=False)
            _formatar_aba(writer.sheets[nome_aba], df_tab)

    output.seek(0)
    return output

# ═══════════════════════════════════════════════════════════════════════════════
# INTERFACE STREAMLIT
# ═══════════════════════════════════════════════════════════════════════════════

def montar_tabela_final(df: pd.DataFrame) -> pd.DataFrame:
    """Reorganiza e renomeia colunas para a tabela final de exibição."""
    MAPA = [
        ("codigo",               "Código"),
        ("descricao",            "Descrição"),
        ("qtd_programada",       "Quantidade Vendida"),
        ("estoque_disponivel",   "Estoque Atual"),
        ("falta_atual",          "Falta Atual"),
        ("quarentena_total",     "Quarentena Total"),
        ("proximas_liberacoes",  "Próximas Liberações"),
        ("quarentena_total",     "Quarentena Disponível"),   # V1 = igual ao total
        ("cobertura_futura",     "Cobertura Futura"),
        ("producao_prevista",    "Produção Prevista"),
        ("data_producao",        "Data Produção"),
        ("Status Operacional",   "Status Operacional"),
        ("alerta_comercial",     "Alerta Comercial"),
        ("percentual_vendido",   "Percentual Vendido x Planejado"),
        ("excesso_vendido",      "Excesso Vendido"),
        ("observacao_planejamento", "Observação Planejamento"),
        ("Observação Sistema",   "Observação Sistema"),
    ]

    df_final = pd.DataFrame()
    visto    = set()
    for col_orig, col_dest in MAPA:
        if col_dest in visto:
            continue
        visto.add(col_dest)
        df_final[col_dest] = df[col_orig].values if col_orig in df.columns else ""

    df_final["Corte Manual"] = ""   # sempre vazia na V1
    return df_final


def main():
    st.set_page_config(
        page_title="Torre de Controle de Estoque",
        page_icon="🏭",
        layout="wide",
    )

    # ── Cabeçalho ─────────────────────────────────────────────────────────────
    st.title("🏭 Torre de Controle de Estoque e Cobertura")
    st.caption("Filial CDSC · Análise de falta de estoque, quarentena e cobertura futura")
    st.divider()

    # ── Upload das planilhas ───────────────────────────────────────────────────
    col1, col2 = st.columns(2)
    with col1:
        f_dem = st.file_uploader(
            "📊 Análise de Demanda — Carteira de Pedidos",
            type=["xlsx"], key="dem",
            help="Aba: CDSC - BLOQ ITEM | Cabeçalho na linha 22"
        )
        f_qua = st.file_uploader(
            "🔒 DW Quarentena WMS",
            type=["xlsx"], key="qua",
            help="Aba: DW_ESTOQUE_QUARENTENA (2) | Enviar sem filtros"
        )
    with col2:
        f_pla = st.file_uploader(
            "📈 Planejado x Vendido",
            type=["xlsx"], key="pla",
            help="Aba: PLAN VENDA VS VENDIDO - CDSC"
        )
        f_ate = st.file_uploader(
            "📋 Planilha de Atendimento",
            type=["xlsx"], key="ate",
            help="Aba: CDSC"
        )

    # Status dos uploads
    uploads_ok = [f_dem, f_qua, f_pla, f_ate]
    n_ok = sum(1 for f in uploads_ok if f is not None)
    if n_ok > 0:
        st.info(f"✅ {n_ok}/4 planilhas carregadas. "
                + ("" if n_ok < 4 else "Pronto para gerar análise!"))

    st.divider()

    # ── Botão principal ────────────────────────────────────────────────────────
    if st.button("⚙️ Gerar Análise", type="primary", use_container_width=True):
        if not f_dem:
            st.error("A planilha **Análise de Demanda** é obrigatória.")
            st.stop()

        with st.spinner("🔄 Lendo bases e processando dados..."):

            # 1. Leitura
            df_dem = read_demanda(f_dem)
            df_qua = read_quarentena(f_qua)        if f_qua else pd.DataFrame()
            df_pla = read_planejado_vendido(f_pla) if f_pla else pd.DataFrame()
            df_ate = read_atendimento(f_ate)       if f_ate else pd.DataFrame()

            if df_dem.empty:
                st.error("Não foi possível ler a base de demanda. Verifique o arquivo.")
                st.stop()

            # 2. Identifica faltantes
            df = identificar_itens_faltantes(df_dem)
            if df.empty:
                st.warning("⚠️ Nenhum item com falta de estoque identificado.")
                st.stop()

            # 3. Enriquece com quarentena, planejado e atendimento
            df = calcular_quarentena(df, df_qua)
            df = consultar_planejado_vendido(df, df_pla)
            df = consultar_atendimento(df, df_ate)

            # 4. Status e observação
            df["Status Operacional"] = df.apply(definir_status_operacional, axis=1)
            df["Observação Sistema"] = df.apply(gerar_observacao_sistema, axis=1)

            # 5. Tabela final
            df_final = montar_tabela_final(df)

        # ── Cards de resumo ────────────────────────────────────────────────────
        st.subheader("📌 Resumo da Análise")
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("📦 Itens com Falta",           len(df_final))
        c2.metric("✅ Cob. por Quarentena",
                   (df_final["Status Operacional"] == "COBERTURA VIA QUARENTENA").sum())
        c3.metric("🟡 Cobertura Parcial",
                   (df_final["Status Operacional"] == "COBERTURA PARCIAL").sum())
        c4.metric("🔴 Críticos",
                   (df_final["Status Operacional"] == "CRÍTICO").sum())
        c5.metric("📈 Venda Acima do Planejado",
                   (df_final["Alerta Comercial"] != "").sum())

        st.divider()

        # ── Filtros ────────────────────────────────────────────────────────────
        st.subheader("🔍 Filtros")
        fc1, fc2, fc3 = st.columns(3)

        with fc1:
            opts_status  = ["Todos"] + sorted(df_final["Status Operacional"].dropna().unique().tolist())
            filtro_status = st.selectbox("Status Operacional", opts_status)

        with fc2:
            filtro_alerta = st.selectbox("Alerta Comercial", ["Todos", "Com Alerta", "Sem Alerta"])

        with fc3:
            filtro_cod = st.text_input("🔎 Código do Produto (busca parcial)")

        dv = df_final.copy()
        if filtro_status != "Todos":
            dv = dv[dv["Status Operacional"] == filtro_status]
        if filtro_alerta == "Com Alerta":
            dv = dv[dv["Alerta Comercial"] != ""]
        elif filtro_alerta == "Sem Alerta":
            dv = dv[dv["Alerta Comercial"] == ""]
        if filtro_cod.strip():
            dv = dv[dv["Código"].astype(str).str.contains(filtro_cod.strip(), case=False, na=False)]

        # ── Tabela principal ───────────────────────────────────────────────────
        st.subheader(f"📋 Relatório Final — {len(dv)} iten(s)")

        styled = dv.style.applymap(
            lambda v: CORES_STATUS_CSS.get(v, ""),
            subset=["Status Operacional"]
        )
        st.dataframe(styled, use_container_width=True, height=520)

        # ── Download ───────────────────────────────────────────────────────────
        st.divider()
        excel_bytes = gerar_excel(df_final)
        st.download_button(
            label="📥 Baixar Relatório Excel (4 abas)",
            data=excel_bytes,
            file_name="torre_controle_cdsc.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )

        # Aviso sobre coluna de corte
        st.caption(
            "💡 A coluna **Corte Manual** está em branco — "
            "preencha diretamente no Excel após o download."
        )


if __name__ == "__main__":
    main()
